import pandas as pd
from typing import List, Dict, Any, Optional
import logging
from datetime import datetime, timedelta
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference
from ..utils.config import Config

logger = logging.getLogger(__name__)

class ExcelExporter:
    """Export price tracking data to Excel files"""
    
    def __init__(self, config: Config = None):
        self.config = config or Config()
        self.output_dir = "exports"
        os.makedirs(self.output_dir, exist_ok=True)
    
    def export_products(self, products: List[Dict[str, Any]], filename: str = None) -> Optional[str]:
        """Export products data to Excel file"""
        
        if not products:
            logger.info("No products to export")
            return None
        
        try:
            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"products_export_{timestamp}.xlsx"
            
            filepath = os.path.join(self.output_dir, filename)
            
            # Create DataFrame
            df = pd.DataFrame(products)
            
            # Reorder columns for better presentation
            preferred_columns = [
                'id', 'title', 'platform', 'current_price', 'target_price',
                'availability', 'rating', 'review_count', 'seller', 'brand',
                'category', 'last_checked', 'created_at', 'url'
            ]
            
            # Keep only existing columns in preferred order
            columns = [col for col in preferred_columns if col in df.columns]
            remaining_columns = [col for col in df.columns if col not in columns]
            final_columns = columns + remaining_columns
            
            df = df[final_columns]
            
            # Create Excel writer with xlsxwriter engine for formatting
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Write main products sheet
                df.to_excel(writer, sheet_name='Products', index=False)
                
                # Get workbook and worksheet for formatting
                workbook = writer.book
                worksheet = writer.sheets['Products']
                
                # Format headers
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Add summary sheet
                self._create_summary_sheet(workbook, products)
            
            logger.info(f"Successfully exported {len(products)} products to {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Failed to export products to Excel: {e}")
            return None
    
    def export_price_history(self, history_data: List[Dict[str, Any]], filename: str = None) -> Optional[str]:
        """Export price history data to Excel file"""
        
        if not history_data:
            logger.info("No price history to export")
            return None
        
        try:
            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"price_history_{timestamp}.xlsx"
            
            filepath = os.path.join(self.output_dir, filename)
            
            # Create DataFrame
            df = pd.DataFrame(history_data)
            
            # Sort by product_id and timestamp
            if 'product_id' in df.columns and 'timestamp' in df.columns:
                df = df.sort_values(['product_id', 'timestamp'])
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Write main history sheet
                df.to_excel(writer, sheet_name='Price History', index=False)
                
                # Format the sheet
                workbook = writer.book
                worksheet = writer.sheets['Price History']
                
                # Format headers
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Create price trend charts for each product
                self._create_price_charts(workbook, df)
            
            logger.info(f"Successfully exported {len(history_data)} price history records to {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Failed to export price history to Excel: {e}")
            return None
    
    def export_comprehensive_report(self, products: List[Dict[str, Any]], 
                                  history_data: List[Dict[str, Any]], 
                                  filename: str = None) -> Optional[str]:
        """Export comprehensive report with products, history, and analysis"""
        
        try:
            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"price_tracker_report_{timestamp}.xlsx"
            
            filepath = os.path.join(self.output_dir, filename)
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Export products if available
                if products:
                    df_products = pd.DataFrame(products)
                    df_products.to_excel(writer, sheet_name='Products', index=False)
                    self._format_sheet(writer.sheets['Products'], "366092")
                
                # Export price history if available
                if history_data:
                    df_history = pd.DataFrame(history_data)
                    df_history = df_history.sort_values(['product_id', 'timestamp']) if 'product_id' in df_history.columns else df_history
                    df_history.to_excel(writer, sheet_name='Price History', index=False)
                    self._format_sheet(writer.sheets['Price History'], "C55A11")
                
                # Create analysis sheets
                workbook = writer.book
                
                if products:
                    self._create_summary_sheet(workbook, products)
                    self._create_price_analysis_sheet(workbook, products, history_data)
                
                if history_data:
                    self._create_trend_analysis_sheet(workbook, history_data)
            
            logger.info(f"Successfully created comprehensive report: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Failed to create comprehensive report: {e}")
            return None
    
    def _format_sheet(self, worksheet, header_color: str):
        """Apply standard formatting to a worksheet"""
        
        # Format headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_summary_sheet(self, workbook, products: List[Dict[str, Any]]):
        """Create a summary analysis sheet"""
        
        summary_sheet = workbook.create_sheet("Summary")
        
        # Calculate metrics
        total_products = len(products)
        active_products = len([p for p in products if p.get('is_active', True)])
        in_stock = len([p for p in products if p.get('availability', False)])
        out_of_stock = total_products - in_stock
        
        # Price analysis
        prices = [p.get('current_price') for p in products if p.get('current_price')]
        avg_price = sum(prices) / len(prices) if prices else 0
        min_price = min(prices) if prices else 0
        max_price = max(prices) if prices else 0
        
        # Platform breakdown
        platforms = {}
        for product in products:
            platform = product.get('platform', 'Unknown')
            platforms[platform] = platforms.get(platform, 0) + 1
        
        # Write summary data
        summary_data = [
            ["Smart Price Tracker Summary Report", ""],
            ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["", ""],
            ["Overall Metrics", ""],
            ["Total Products", total_products],
            ["Active Products", active_products],
            ["In Stock", in_stock],
            ["Out of Stock", out_of_stock],
            ["", ""],
            ["Price Analysis", ""],
            ["Average Price", f"${avg_price:.2f}" if avg_price > 0 else "N/A"],
            ["Minimum Price", f"${min_price:.2f}" if min_price > 0 else "N/A"],
            ["Maximum Price", f"${max_price:.2f}" if max_price > 0 else "N/A"],
            ["", ""],
            ["Platform Breakdown", ""],
        ]
        
        # Add platform data
        for platform, count in platforms.items():
            summary_data.append([f"{platform.title()}", count])
        
        # Write to sheet
        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                cell = summary_sheet.cell(row=row_idx, column=col_idx, value=cell_value)
                
                # Format headers and titles
                if row_idx == 1:  # Main title
                    cell.font = Font(bold=True, size=16, color="FFFFFF")
                    cell.fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
                elif col_idx == 1 and cell_value and any(keyword in str(cell_value) for keyword in ["Metrics", "Analysis", "Breakdown"]):
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        
        # Auto-adjust column widths
        summary_sheet.column_dimensions['A'].width = 25
        summary_sheet.column_dimensions['B'].width = 15
    
    def _create_price_analysis_sheet(self, workbook, products: List[Dict[str, Any]], history_data: List[Dict[str, Any]]):
        """Create price analysis sheet with trends and insights"""
        
        analysis_sheet = workbook.create_sheet("Price Analysis")
        
        # Price change analysis
        if history_data:
            df_history = pd.DataFrame(history_data)
            
            # Group by product and calculate price changes
            price_changes = []
            
            if 'product_id' in df_history.columns and 'price' in df_history.columns:
                for product_id in df_history['product_id'].unique():
                    product_history = df_history[df_history['product_id'] == product_id].sort_values('timestamp')
                    
                    if len(product_history) >= 2:
                        first_price = product_history.iloc[0]['price']
                        last_price = product_history.iloc[-1]['price']
                        
                        if first_price and last_price and first_price > 0:
                            change_pct = ((last_price - first_price) / first_price) * 100
                            
                            # Find product title
                            product_title = "Unknown"
                            for product in products:
                                if product.get('id') == product_id:
                                    product_title = product.get('title', 'Unknown')[:50]
                                    break
                            
                            price_changes.append({
                                'Product': product_title,
                                'First Price': first_price,
                                'Last Price': last_price,
                                'Change %': change_pct,
                                'Change $': last_price - first_price
                            })
            
            # Sort by percentage change
            price_changes.sort(key=lambda x: x['Change %'], reverse=True)
            
            # Write headers
            headers = ['Product', 'First Price', 'Last Price', 'Change %', 'Change $']
            for col_idx, header in enumerate(headers, 1):
                cell = analysis_sheet.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            
            # Write data
            for row_idx, change_data in enumerate(price_changes, 2):
                analysis_sheet.cell(row=row_idx, column=1, value=change_data['Product'])
                analysis_sheet.cell(row=row_idx, column=2, value=f"${change_data['First Price']:.2f}")
                analysis_sheet.cell(row=row_idx, column=3, value=f"${change_data['Last Price']:.2f}")
                analysis_sheet.cell(row=row_idx, column=4, value=f"{change_data['Change %']:.1f}%")
                analysis_sheet.cell(row=row_idx, column=5, value=f"${change_data['Change $']:.2f}")
        
        # Auto-adjust column widths
        for column in analysis_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 60)
            analysis_sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_trend_analysis_sheet(self, workbook, history_data: List[Dict[str, Any]]):
        """Create trend analysis with charts"""
        
        trend_sheet = workbook.create_sheet("Trends")
        
        # Add some basic trend information
        trend_sheet.cell(row=1, column=1, value="Price Trend Analysis")
        trend_sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
        
        trend_sheet.cell(row=3, column=1, value="Total price records:")
        trend_sheet.cell(row=3, column=2, value=len(history_data))
        
        # Could add more sophisticated trend analysis here
        # For now, this provides a foundation for future enhancements
    
    def _create_price_charts(self, workbook, df_history: pd.DataFrame):
        """Create price trend charts for products"""
        
        # This is a placeholder for chart creation
        # Charts would require more complex implementation with openpyxl
        # For now, the data is available for manual chart creation in Excel
        pass
    
    def get_export_directory(self) -> str:
        """Get the export directory path"""
        return os.path.abspath(self.output_dir)
    
    def list_exports(self) -> List[str]:
        """List all export files in the export directory"""
        try:
            files = [f for f in os.listdir(self.output_dir) if f.endswith('.xlsx')]
            return sorted(files, reverse=True)  # Most recent first
        except Exception as e:
            logger.error(f"Failed to list exports: {e}")
            return [] 